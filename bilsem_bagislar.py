import pandas as pd
import re
import locale
locale.setlocale(locale.LC_ALL, 'tr_TR')
import openpyxl
from openpyxl.formula.translate import Translator

def clear(x):
    non_alpha = r'[^a-zA-ZçğıöşüÇĞİÖŞÜ\u00E7\u011F\u0131\u00F6\u015F\u00FC\u00C7\u011E\u0130\u00D6\u015E\u00DC]'
    ret = x
    ret = re.sub('SN: ','SN:', ret)
    ret = re.sub('SN:\d+','_ ', ret)
    ret = re.sub('Banka: \d+','_ ', ret)
    ret = re.sub('GönBanka:\d+ ','_ ', ret)
    ret = re.sub('GönŞube:\d+ ','_ ', ret)
    ret = re.sub('EftRef:\d+ ','_ ', ret)
    ret = re.sub('Eft Otomatik Muhasebe',' ', ret)
    ret = re.sub('Eft Otomatik M.*?(?=\s|$)','_ ',ret)
    ret = re.sub('Eft O.*?(?=\s|$)','_ ',ret)
    ret = ret.replace('_ ','')
    ret = ret.replace('i','İ').upper()
    """
    ret = ret.replace('Ç','C')
    ret = ret.replace('Ğ','G')
    ret = ret.replace('İ','I')
    ret = ret.replace('Ö','O')
    ret = ret.replace('Ş','S')
    ret = ret.replace('Ü','U')
    """
    ret = re.sub(non_alpha, ' ', ret)
    ret = ret.strip()
    return ret

def cmatch(x, tokens):
    tokens = np.unique(tokens).tolist()
    isim = x.İSİM.split(' ')
    
    anne = clear(x['BABA ADI']).split(' ')
    # anne/baba adı veya soyadı çocuğun ismi içinde olduğu zaman çift sayıyor
    anne = [a for a in anne if not a in isim]
    
    baba = clear(x['ANNE ADI']).split(' ')
    # anne/baba adı veya soyadı çocuğun ismi içinde olduğu zaman çift sayıyor
    baba = [b for b in baba if not b in isim]
    
    m1 = sum([1   for t in tokens for o in isim if t==o])
    m2 = sum([0.3 for t in tokens for o in anne if t==o])
    m3 = sum([0.3 for t in tokens for o in baba if t==o])
    return m1+m2+m3

def aciklama_match(x):
    x = clear(x)
    
    # manuel müdahaleler -------------------------------------
    x = x.replace('KEREM ÇINAR','KEREM')
    # manuel müdahaleler -------------------------------------
    
    tokens = [t for t in x.split(' ') if t!='']

    dff = dof.assign(PUAN=dof.fillna('').apply(lambda x: cmatch(x, tokens), axis=1))
    dff = dff[dff.PUAN>1].sort_values('PUAN', ascending=False)
    max_puan = dff.PUAN.max()
    dff = dff[dff.PUAN==max_puan]
    
    if dff.shape[0]==0:
        oid=-1
    elif dff.shape[0]==1:
        oid=dff.iloc[0,].name
    else:
        oid=-2
    
    ret = pd.Series({'ÖğrenciID': oid ,
                     'EşleşmeAdet': len(dff),
                     'Eşleşenler':dff.reset_index().values,
                    })
    
    return ret

def excelize(do, dff):
    # create file -------------------------------------------------------------
    file_name = 'Bilsem_Bagis.xlsx'
    do.to_excel(file_name, sheet_name='Ogrenciler') # save ogrenciler to file

    # open file ---------------------------------------------------------------
    book = openpyxl.load_workbook(file_name, )
    writer = pd.ExcelWriter(file_name, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    dff.to_excel(writer, "Hareketler") # save Hareketler to file
    writer.save()

    # re-open ogrenciler to add hareket data ---------------------------------
    book = openpyxl.load_workbook(file_name)
    ws = book['Ogrenciler']

    # calculate distinct dates from hareketler -------------------------------
    dates = dff.Dönem.drop_duplicates().sort_values().values.tolist()

    # find last colum to add dates
    max_col_num = ws.max_column + 1
    max_col_cell = ws.cell(1, max_col_num)
    max_col_adr1 = f'{max_col_cell.column_letter}${max_col_cell.row}'
    max_col_adr2 = f'{max_col_cell.column_letter}{max_col_cell.row}'
    
    
    # write calculation formula to first cell --------------------------------
    formula = f'=SUMIFS(Hareketler!$G:$G,Hareketler!$K:$K,Ogrenciler!$A2,Hareketler!$J:$J,Ogrenciler!{max_col_adr1})'
    ws[max_col_adr2] = formula
    ws.cell(do.shape[0]+2,1).value = -1

    # write all dates to columns and copy the formula to all cells -----------
    for i,d in enumerate(dates):
        ws.cell(1, max_col_num+i).value = d
        for r in range(2,do.shape[0]+3):
            c = ws.cell(r, max_col_num+i)
            c.value = Translator(formula, origin="X2").translate_formula(f'{c.column_letter}{c.row}')

            
    # Toplam sütunu -----------------------------------------------------------
    start_col_num = ws.max_column + 1
    
    # Atölye Sütunları
    ekler = ['*', 'KİMYA', 'SERAMİK', 'RESİM', 'MÜZİK', 'ROBOT']
    for i, ek in enumerate(ekler):
        c = ws.cell(1, start_col_num + i)
        c.value = ek
    
    c = ws.cell(1, start_col_num)
    org1 = f'{c.column_letter}${c.row}'
    c = ws.cell(2, start_col_num)
    org2 = f'{c.column_letter}${c.row}'
    
    formula2 = f"""=SUMIFS(Hareketler!$G:$G,Hareketler!$K:$K,Ogrenciler!$A2,Hareketler!$I:$I,"*"&Ogrenciler!{org1}&"*")"""
    c.value = formula2
    
    for i,d in enumerate(ekler):
        for r in range(2,do.shape[0]+3):
            c = ws.cell(r, start_col_num+i)
            fcell = f'{c.column_letter}{c.row}'
            c.value = Translator(formula2, origin=org2.replace('$','')).translate_formula(fcell)
    
    
    # add filter
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws['N2'] 
    
    # save & close file -------------------------------------------------------
    book.save(file_name)
    book.close()

def run():
    # Dosyaları oku -----------------------------------------------------------------------------------------
    df = pd.read_excel('data/Hesap Hareketleri_20220801_20230118.xlsx', skiprows=8)
    do = pd.read_excel('data/Bilsem_12_12_2022 14_07_53.xlsx', sheet_name='genel liste')
    dd = pd.read_excel('data/Bilsem_12_12_2022 14_07_53.xlsx', sheet_name='destek')
    do = pd.merge(left=do, right=dd[['TC NO','İSİM']].assign(DESTEK=1), how='left',
             left_on=['TC NO','İSİM'], right_on=['TC NO','İSİM']
            )

    # Açıklamaları temizle ----------------------------------------------------------------------------------
    dfc = df.assign(Açıklama2=df.Açıklama.apply(lambda x: clear(x)))
    #dfc = dfc.sample(20, random_state=12)
    #dfc = dfc[dfc.Açıklama2.str.contains('KEREM ÇINAR')]


    # Dönem bilgisini ekle, tarih'ten Yıl-Ay formatına çevir ------------------------------------------------
    dfc = dfc.assign(Dönem = lambda x: pd.to_datetime(x.Tarih, format='%d.%m.%Y %H:%M').dt.strftime('%Y-%m'))


    # Açıklamalar ile öğrencileri eşleştir ------------------------------------------------------------------
    dfcc = pd.merge(dfc, dfc.apply(lambda x: aciklama_match(x.Açıklama2), axis=1),
             left_index=True, right_index=True
            )


    # Birden fazla öğrenci ile eşleşen kayıtları paylaştır --------------------------------------------------
    dff = pd.concat(
        [
            dfcc[dfcc.ÖğrenciID!=-2],
            pd.concat([
                dfcc[dfcc.ÖğrenciID==-2].assign(ÖğrenciID=lambda x: x.Eşleşenler.apply(lambda y: y[0][0]))\
                    .assign(**{'Tutar(TRY)': lambda x: x[['Tutar(TRY)']]/2}), 
                dfcc[dfcc.ÖğrenciID==-2].assign(ÖğrenciID=lambda x: x.Eşleşenler.apply(lambda y: y[1][0]))\
                    .assign(**{'Tutar(TRY)': lambda x: x[['Tutar(TRY)']]/2}),     
            ]).sort_index()    ]
    )
    dff    
    
    excelize(do, dff)
